"""Write the YOLO11m training hyperparameters to Excel: parameter and value only."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(r"C:\Users\cihan\Desktop\mercimek_hiperparametreler.xlsx")

PARAMS: list[tuple[str, str]] = [
    ("model", "yolo11m.pt"),
    ("data", "/content/data_local.yaml"),
    ("epochs", "100"),
    ("imgsz", "2048"),
    ("batch", "12"),
    ("patience", "25"),
    ("optimizer", "auto"),
    ("device", "0"),
    ("workers", "16"),
    ("cache", "False"),
    ("amp", "True"),
    ("seed", "17"),
    ("deterministic", "False"),
    ("rect", "False"),
    ("single_cls", "False"),
    ("lr0", "0.01"),
    ("lrf", "0.01"),
    ("cos_lr", "True"),
    ("momentum", "0.937"),
    ("weight_decay", "0.0005"),
    ("warmup_epochs", "3.0"),
    ("warmup_momentum", "0.8"),
    ("warmup_bias_lr", "0.1"),
    ("nbs", "64"),
    ("box", "7.5"),
    ("cls", "0.5"),
    ("dfl", "1.5"),
    ("iou", "0.7"),
    ("max_det", "300"),
    ("mosaic", "1.0"),
    ("close_mosaic", "10"),
    ("degrees", "10.0"),
    ("flipud", "0.5"),
    ("fliplr", "0.5"),
    ("scale", "0.5"),
    ("translate", "0.1"),
    ("shear", "0.0"),
    ("perspective", "0.0"),
    ("hsv_h", "0.015"),
    ("hsv_s", "0.7"),
    ("hsv_v", "0.4"),
    ("erasing", "0.4"),
    ("mixup", "0.0"),
    ("cutmix", "0.0"),
    ("copy_paste", "0.0"),
]

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def main() -> int:
    wb = Workbook()
    ws = wb.active
    ws.title = "Hiperparametreler"

    for col, name in enumerate(("Parametre", "Değer"), start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor="2E5C9A")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    ws.row_dimensions[1].height = 20

    for index, (param, value) in enumerate(PARAMS, start=2):
        left = ws.cell(row=index, column=1, value=param)
        left.font = Font(name="Consolas", size=10)
        left.border = BORDER

        right = ws.cell(row=index, column=2, value=value)
        right.font = Font(name="Consolas", size=10)
        right.alignment = Alignment(horizontal="left")
        right.border = BORDER

        if index % 2 == 0:
            fill = PatternFill("solid", fgColor="F5F7FA")
            left.fill = fill
            right.fill = fill

    ws.column_dimensions[get_column_letter(1)].width = 22
    ws.column_dimensions[get_column_letter(2)].width = 32
    ws.freeze_panes = ws.cell(row=2, column=1)

    wb.save(OUT)
    print(f"Yazildi: {OUT}  ({len(PARAMS)} parametre)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
